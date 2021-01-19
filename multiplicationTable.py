import openpyxl
from openpyxl.styles import Font
#import pyinputplus as pyip

#N = int(sys.argv[2])
#N = pyip.inputInt(prompt='Input an int (N): ')
N = 26

wb = openpyxl.Workbook()
sheet = wb.active
fontObjBold = Font(size=15,bold=True)
#create headers on top row and leftmost column
for i in range(N):
    sheet.cell(row = 1,column=i+2).value = i+1 #top row
    sheet.cell(row=i+2,column=1).value=i+1 #leftmost column
    sheet.cell(row = 1,column=i+2).font = fontObjBold #top row
    sheet.cell(row=i+2,column=1).font=fontObjBold #leftmost column

#populate multiplication table
for row in range(N):
    for col  in range(N):
        sheet.cell(row=row+2, column=col+2).value = (row+1)*(col+1)

wb.save('multiplicationTable.xlsx')


















