#! python3
#updateProduce.py - Corrects costs in produce sales spreadsheet.
import openpyxl, os
#os.chdir('')

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

#produce types and their updated prices
PRICE_UPDATES = {'Garlic':3.07,
        'Celery':1.19,
        'Lemon': 1.27 }
#loop through rows and update price
for rowNum in range(2,sheet.max_row): #skip first row, b/c that is header
    produceName = sheet.cell(row=rowNum,column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('updateProduceSales.xlsx')









































